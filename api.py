"""
FastAPI server — exposes QA Intelligence backend as a REST API.
Run: uvicorn backend.api:app --reload --port 8000
"""

from __future__ import annotations

import os
import sys
import json
import uuid
import tempfile
import threading
import concurrent.futures
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, File, Form, Request, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel

sys.path.insert(0, str(Path(__file__).parent))

from config import settings
from ingestion.document_loader import document_loader
from ingestion.chunker import chunker
from rag_engine.vector_store import vector_store
from graph_builder.entity_extractor import entity_extractor
from graph_builder.relationship_builder import relationship_builder
from graph_builder.neo4j_client import get_graph
from graph_builder.graph_queries import GraphQueryEngine
from orchestrator.qa_pipeline import qa_pipeline
from test_generator.llm_client import llm_client
from project_manager import project_manager

app = FastAPI(title="QA Intelligence API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response: Response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    if request.url.scheme == "https":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

_graph_lock = threading.Lock()

# ── Pricing table ────────────────────────────────────────────────────────────
_PRICING: Dict[str, tuple] = {
    "meta-llama/Llama-3.3-70B-Instruct":       (0.23, 0.40),
    "meta-llama/Meta-Llama-3.1-8B-Instruct":   (0.06, 0.06),
    "meta-llama/Meta-Llama-3.1-70B-Instruct":  (0.52, 0.75),
    "meta-llama/Meta-Llama-3.1-405B-Instruct": (2.70, 2.70),
    "mistralai/Mixtral-8x7B-Instruct-v0.1":    (0.27, 0.27),
    "deepseek-ai/DeepSeek-R1":                 (0.55, 2.19),
    "Qwen/Qwen2.5-72B-Instruct":               (0.35, 0.40),
    "gpt-4o":                                  (2.50, 10.00),
    "gpt-4o-mini":                             (0.15, 0.60),
    "gpt-4-turbo":                             (10.00, 30.00),
    "gpt-3.5-turbo":                           (0.50, 1.50),
    "llama-3.3-70b-versatile":                 (0.59, 0.79),
    "llama-3.1-8b-instant":                    (0.05, 0.08),
    "mixtral-8x7b-32768":                      (0.24, 0.24),
    "gemma2-9b-it":                            (0.20, 0.20),
}

# ── Providers ─────────────────────────────────────────────────────────────────
_PROVIDERS: Dict[str, Any] = {
    "deepinfra": {
        "name": "DeepInfra",
        "base_url": "https://api.deepinfra.com/v1/openai",
        "key_required": True,
        "models": [
            "meta-llama/Llama-3.3-70B-Instruct",
            "meta-llama/Meta-Llama-3.1-70B-Instruct",
            "meta-llama/Meta-Llama-3.1-8B-Instruct",
            "mistralai/Mixtral-8x7B-Instruct-v0.1",
            "Qwen/Qwen2.5-72B-Instruct",
        ],
    },
    "openai": {
        "name": "OpenAI",
        "base_url": "https://api.openai.com/v1",
        "key_required": True,
        "models": ["gpt-4o", "gpt-4o-mini"],
    },
    "groq": {
        "name": "Groq",
        "base_url": "https://api.groq.com/openai/v1",
        "key_required": True,
        "models": [
            "llama-3.3-70b-versatile",
            "llama-3.1-8b-instant",
            "gemma2-9b-it",
        ],
    },
    "ollama": {
        "name": "Ollama (Local)",
        "base_url": "http://localhost:11434/v1",
        "key_required": False,
        "models": ["llama3.2", "llama3.1", "mistral", "codellama", "phi3"],
    },
}

# ── Runtime LLM config (updated via /api/settings/llm) ────────────────────────
_runtime_llm: Dict[str, str] = {
    "provider": "deepinfra",
    "model": settings.LLM_MODEL,
}


# ── Models ───────────────────────────────────────────────────────────────────

class AnalyzeRequest(BaseModel):
    user_story: str
    clarifying_answers: Optional[Dict[str, str]] = None


class ClarifyRequest(BaseModel):
    user_story: str


class ChatRequest(BaseModel):
    message: str
    history: List[Dict[str, str]] = []
    context: str = ""


# ── Helpers ──────────────────────────────────────────────────────────────────

def _ingest_file(file_bytes: bytes, filename: str, doc_type: str) -> dict:
    result = {"file": filename, "type": doc_type, "status": "PENDING",
              "chunks": 0, "entities": 0, "rels": 0, "steps": []}
    try:
        suffix = "." + filename.split(".")[-1] if "." in filename else ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        source_id = filename.rsplit(".", 1)[0]
        document = document_loader.load_file(tmp_path, doc_type)
        document["source_id"] = source_id
        os.unlink(tmp_path)

        relevance = llm_client.check_document_relevance(
            content=document.get("content", ""), filename=filename, doc_type=doc_type
        )
        if not relevance["is_relevant"]:
            result["status"] = "REJECTED"
            result["reason"] = relevance["reason"]
            return result

        result["detected_type"] = relevance.get("detected_type", doc_type)

        chunks = chunker.chunk(document)
        chunks_stored = vector_store.add_chunks(chunks)
        result["chunks"] = chunks_stored

        try:
            extraction = entity_extractor.extract(document)
            with _graph_lock:
                gr = relationship_builder.ingest(extraction, source_id)
            result["entities"] = gr["entities"]
            result["rels"] = gr["relationships"]
        except Exception:
            pass

        result["status"] = "OK"
    except Exception as e:
        result["status"] = "FAIL"
        result["error"] = str(e)
    return result


# ── Analysis History Manager ─────────────────────────────────────────────────

class AnalysisHistoryManager:
    """Persists per-project analysis history.

    Uses MongoDB when MONGODB_URI is configured; falls back to JSON files on disk.
    """

    def __init__(self):
        from db.mongo_client import get_db
        db = get_db()
        if db is not None:
            from db.repositories import AnalysisRepository
            self._repo = AnalysisRepository(db["analyses"])
            self._mode = "mongo"
        else:
            self._mode = "file"
            self._lock = threading.Lock()
            base = Path(settings.DATA_DIR) if settings.DATA_DIR else Path(__file__).parent / "data"
            self._dir = base / "analyses"
            self._dir.mkdir(parents=True, exist_ok=True)

    # ── file helpers ──────────────────────────────────────────────────────────

    def _fpath(self, project_id: str) -> Path:
        return self._dir / f"{project_id}.json"

    def _fload(self, project_id: str) -> List[Dict]:
        p = self._fpath(project_id)
        if not p.exists():
            return []
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return []

    def _fsave(self, project_id: str, entries: List[Dict]) -> None:
        self._fpath(project_id).write_text(
            json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    # ── public API ────────────────────────────────────────────────────────────

    def save(self, project_id: str, story: str, result: Dict) -> Dict:
        if self._mode == "mongo":
            return self._repo.save(project_id, story, result)
        entry = {
            "id": uuid.uuid4().hex[:12],
            "story": story,
            "result": result,
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        with self._lock:
            entries = self._fload(project_id)
            entries.insert(0, entry)
            self._fsave(project_id, entries)
        return entry

    def list(self, project_id: str) -> List[Dict]:
        if self._mode == "mongo":
            return self._repo.list(project_id)
        with self._lock:
            return self._fload(project_id)

    def delete_one(self, project_id: str, analysis_id: str) -> bool:
        if self._mode == "mongo":
            return self._repo.delete_one(project_id, analysis_id)
        with self._lock:
            entries = self._fload(project_id)
            new_entries = [e for e in entries if e["id"] != analysis_id]
            if len(new_entries) == len(entries):
                return False
            self._fsave(project_id, new_entries)
            return True

    def delete_all(self, project_id: str) -> None:
        if self._mode == "mongo":
            self._repo.delete_all(project_id)
            return
        with self._lock:
            p = self._fpath(project_id)
            if p.exists():
                p.unlink()


analysis_history = AnalysisHistoryManager()


# ── Ingested Document Tracker ────────────────────────────────────────────────

class DocumentTracker:
    """Tracks ingested document metadata in MongoDB (no-op when MongoDB is off)."""

    def __init__(self):
        from db.mongo_client import get_db
        db = get_db()
        if db is not None:
            from db.repositories import DocumentRepository
            self._repo = DocumentRepository(db["ingested_documents"])
            self._enabled = True
        else:
            self._enabled = False

    def save(self, project_id: str, filename: str, doc_type: str, ingest_result: dict) -> Optional[dict]:
        if not self._enabled:
            return None
        return self._repo.save(project_id, filename, doc_type, ingest_result)

    def list(self, project_id: str) -> List[dict]:
        if not self._enabled:
            return []
        return self._repo.list(project_id)

    def delete_all(self, project_id: str) -> None:
        if self._enabled:
            self._repo.delete_all(project_id)


document_tracker = DocumentTracker()


# ── Routes ───────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/settings")
def get_settings():
    model = _runtime_llm["model"]
    price = _PRICING.get(model)
    usage = llm_client.get_usage()
    inp_tok = usage["input_tokens"]
    out_tok = usage["output_tokens"]
    cost = None
    if price and (inp_tok or out_tok):
        cost = {
            "input_cost": round(inp_tok / 1_000_000 * price[0], 6),
            "output_cost": round(out_tok / 1_000_000 * price[1], 6),
            "total_cost": round(inp_tok / 1_000_000 * price[0] + out_tok / 1_000_000 * price[1], 6),
            "input_rate": price[0],
            "output_rate": price[1],
        }
    return {
        "model": model,
        "provider": _runtime_llm["provider"],
        "embed_model": settings.EMBED_MODEL,
        "entity_model": settings.ENTITY_MODEL,
        "usage": {"input_tokens": inp_tok, "output_tokens": out_tok},
        "cost": cost,
        "api_key_set": llm_client.is_api_key_set(),
    }


@app.get("/api/settings/providers")
def get_providers():
    return _PROVIDERS


class LLMConfigRequest(BaseModel):
    provider: str
    model: str
    api_key: str = ""


@app.post("/api/settings/llm")
def update_llm_settings(req: LLMConfigRequest):
    provider = _PROVIDERS.get(req.provider)
    if not provider:
        raise HTTPException(status_code=400, detail=f"Unknown provider: {req.provider}")
    if provider["key_required"] and not req.api_key:
        raise HTTPException(status_code=400, detail="API key is required for this provider")
    llm_client.configure(
        model=req.model,
        api_key=req.api_key or "ollama",
        base_url=provider["base_url"],
    )
    _runtime_llm["provider"] = req.provider
    _runtime_llm["model"] = req.model
    return {"status": "ok", "provider": req.provider, "model": req.model}


@app.post("/api/settings/reset-usage")
def reset_usage():
    llm_client.reset_usage()
    return {"status": "ok"}


@app.get("/api/kb/status")
def kb_status():
    try:
        gq = GraphQueryEngine(get_graph())
        graph_stats = gq.get_graph_stats()
    except Exception as e:
        graph_stats = {"error": str(e)}
    return {
        "chunks": vector_store.count(),
        "sources": vector_store.get_all_sources(),
        "graph": graph_stats,
    }


@app.delete("/api/kb")
def clear_kb():
    try:
        vector_store.clear()
        get_graph().clear()
        return {"status": "cleared"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ingest")
async def ingest(
    files: List[UploadFile] = File(...),
    doc_type: str = Form("auto"),
):
    # Read all file bytes upfront (async) before handing off to threads
    file_payloads = [(f.filename, await f.read()) for f in files]
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [
            executor.submit(_ingest_file, data, filename, doc_type)
            for filename, data in file_payloads
        ]
        results = [f.result() for f in futures]
    return {"results": results}


@app.post("/api/analyze")
def analyze(body: AnalyzeRequest):
    if not body.user_story.strip():
        raise HTTPException(status_code=400, detail="user_story is required")
    try:
        result = qa_pipeline.run(user_story=body.user_story)
        # Snapshot token usage into the response
        result["token_usage"] = llm_client.get_usage()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _ingest_file_to_store(file_bytes: bytes, filename: str, doc_type: str, store) -> dict:
    result = {"file": filename, "type": doc_type, "status": "PENDING",
              "chunks": 0, "entities": 0, "rels": 0, "steps": []}
    try:
        suffix = "." + filename.split(".")[-1] if "." in filename else ".txt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        source_id = filename.rsplit(".", 1)[0]
        document = document_loader.load_file(tmp_path, doc_type)
        document["source_id"] = source_id
        os.unlink(tmp_path)

        relevance = llm_client.check_document_relevance(
            content=document.get("content", ""), filename=filename, doc_type=doc_type
        )
        if not relevance["is_relevant"]:
            result["status"] = "REJECTED"
            result["reason"] = relevance["reason"]
            return result

        result["detected_type"] = relevance.get("detected_type", doc_type)
        chunks = chunker.chunk(document)
        chunks_stored = store.add_chunks(chunks)
        result["chunks"] = chunks_stored

        try:
            extraction = entity_extractor.extract(document)
            with _graph_lock:
                gr = relationship_builder.ingest(extraction, source_id)
            result["entities"] = gr["entities"]
            result["rels"] = gr["relationships"]
        except Exception:
            pass

        result["status"] = "OK"
    except Exception as e:
        result["status"] = "FAIL"
        result["error"] = str(e)
    return result


# ── Project endpoints ─────────────────────────────────────────────────────────

class CreateProjectRequest(BaseModel):
    name: str
    description: str = ""

@app.get("/api/projects")
def list_projects():
    projects = project_manager.list_all()
    # Enrich each project with kb stats
    enriched = []
    for p in projects:
        try:
            from rag_engine.vector_store import get_vector_store
            vs = get_vector_store(p["id"])
            chunks = vs.count()
            sources = vs.get_all_sources()
        except Exception:
            chunks = 0
            sources = []
        enriched.append({**p, "chunks": chunks, "file_count": len(sources)})
    return {"projects": enriched}

@app.post("/api/projects")
def create_project(body: CreateProjectRequest):
    project = project_manager.create(body.name, body.description)
    return project

@app.get("/api/projects/{project_id}")
def get_project(project_id: str):
    project = project_manager.get(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    from rag_engine.vector_store import get_vector_store
    vs = get_vector_store(project_id)
    return {**project, "chunks": vs.count(), "sources": vs.get_all_sources()}

@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str):
    # Clear KB
    try:
        from rag_engine.vector_store import get_vector_store
        vs = get_vector_store(project_id)
        vs.clear()
    except Exception:
        pass
    # Clear MongoDB-backed data for this project
    analysis_history.delete_all(project_id)
    document_tracker.delete_all(project_id)
    ok = project_manager.delete(project_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"status": "deleted"}

@app.get("/api/projects/{project_id}/kb/status")
def project_kb_status(project_id: str):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    from rag_engine.vector_store import get_vector_store
    vs = get_vector_store(project_id)
    try:
        gq = GraphQueryEngine(get_graph(project_id))
        graph_stats = gq.get_graph_stats()
    except Exception as e:
        graph_stats = {"error": str(e)}
    return {
        "chunks": vs.count(),
        "sources": vs.get_all_sources(),
        "graph": graph_stats,
    }

@app.delete("/api/projects/{project_id}/kb")
def clear_project_kb(project_id: str):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    try:
        from rag_engine.vector_store import get_vector_store
        vs = get_vector_store(project_id)
        vs.clear()
        get_graph(project_id).clear()
        return {"status": "cleared"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

_MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20 MB per file


@app.post("/api/projects/{project_id}/ingest")
async def project_ingest(
    project_id: str,
    files: List[UploadFile] = File(...),
    doc_type: str = Form("auto"),
):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    from rag_engine.vector_store import get_vector_store
    project_store = get_vector_store(project_id)
    file_payloads = []
    for f in files:
        data = await f.read()
        if len(data) > _MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail=f"{f.filename} exceeds 20 MB limit")
        file_payloads.append((f.filename, data))
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [
            executor.submit(_ingest_file_to_store, data, filename, doc_type, project_store)
            for filename, data in file_payloads
        ]
        results = [f.result() for f in futures]
    # Persist document metadata to MongoDB (no-op if MongoDB is not configured)
    for (filename, _), result in zip(file_payloads, results):
        document_tracker.save(project_id, filename, doc_type, result)
    return {"results": results}


@app.get("/api/projects/{project_id}/documents")
def list_project_documents(project_id: str):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"documents": document_tracker.list(project_id)}

import re as _re

_GREETING_RE = _re.compile(
    r"^\s*(hi+|hey+|hello+|helo+|howdy|good\s*(morning|afternoon|evening|day|night)|"
    r"how\s+(are\s+you|r\s+u)|what'?s\s+up|sup|thanks?(\s+you)?|thank\s+you|"
    r"ty|thx|okay|ok|got\s+it|great|nice|cool|sure|alright|welcome)\s*[!.?]*\s*$",
    _re.IGNORECASE,
)
def _validate_story(text: str) -> None:
    """Raise 400 if the input is a greeting, too short, or not a user story."""
    stripped = text.strip()
    if _GREETING_RE.match(stripped):
        raise HTTPException(
            status_code=400,
            detail="Please enter a user story or feature description, not a greeting."
        )


@app.post("/api/projects/{project_id}/clarify")
def project_clarify(project_id: str, body: ClarifyRequest):
    if not body.user_story.strip():
        raise HTTPException(status_code=400, detail="user_story is required")
    _validate_story(body.user_story)
    try:
        from rag_engine.vector_store import get_vector_store
        from rag_engine.retriever import retriever
        store = get_vector_store(project_id)
        rag_result = retriever.retrieve(body.user_story, store=store, top_k=5)
        rag_context = retriever.build_context_string(rag_result, max_chars=2000)
        questions = llm_client.generate_clarification_questions(body.user_story, rag_context)
        return {"questions": questions}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/projects/{project_id}/chat")
def project_chat(project_id: str, body: ChatRequest):
    if not body.message.strip():
        raise HTTPException(status_code=400, detail="message is required")

    def token_stream():
        llm_client.set_project_context(project_id)
        try:
            for token in llm_client.chat_stream(
                message=body.message,
                history=body.history,
                context=body.context,
            ):
                yield token
        except Exception as e:
            yield f"\n[Error: {str(e)}]"
        finally:
            llm_client.clear_project_context()

    return StreamingResponse(token_stream(), media_type="text/plain")


@app.post("/api/projects/{project_id}/analyze")
def project_analyze(project_id: str, body: AnalyzeRequest):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    if not body.user_story.strip():
        raise HTTPException(status_code=400, detail="user_story is required")
    _validate_story(body.user_story)
    llm_client.set_project_context(project_id)
    try:
        result = qa_pipeline.run(
            user_story=body.user_story,
            project_id=project_id,
            clarifying_answers=body.clarifying_answers,
        )
        result["token_usage"] = llm_client.get_project_usage(project_id)
        # Auto-persist to analysis history
        entry = analysis_history.save(project_id, body.user_story, result)
        result["_history_id"] = entry["id"]
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        llm_client.clear_project_context()


# ── Analysis History endpoints ────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/analyses")
def list_analyses(project_id: str):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"analyses": analysis_history.list(project_id)}


@app.delete("/api/projects/{project_id}/analyses")
def clear_analyses(project_id: str):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    analysis_history.delete_all(project_id)
    return {"status": "cleared"}


@app.delete("/api/projects/{project_id}/analyses/{analysis_id}")
def delete_analysis(project_id: str, analysis_id: str):
    if not project_manager.get(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    ok = analysis_history.delete_one(project_id, analysis_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Analysis not found")
    return {"status": "deleted"}


# ── Export analysis as Excel ──────────────────────────────────────────────────

class ExportRequest(BaseModel):
    result: Dict[str, Any]
    feature_name: Optional[str] = None

@app.post("/api/projects/{project_id}/export/excel")
def export_excel(project_id: str, body: ExportRequest):
    """Generate a multi-sheet Excel workbook from an analysis result."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse as SR

    r = body.result
    feature = body.feature_name or r.get("feature_name") or r.get("detected_module") or "Analysis"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Style helpers ─────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="4F46E5")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FILL  = PatternFill("solid", fgColor="EEF2FF")
    SECTION_FONT  = Font(bold=True, color="3730A3", size=10)
    WARNING_FILL  = PatternFill("solid", fgColor="FFFBEB")
    P1_FILL       = PatternFill("solid", fgColor="FEF2F2")
    P2_FILL       = PatternFill("solid", fgColor="FFF7ED")
    P3_FILL       = PatternFill("solid", fgColor="FEFCE8")
    P4_FILL       = PatternFill("solid", fgColor="F0FDF4")
    HIGH_FILL     = PatternFill("solid", fgColor="FEF2F2")
    MED_FILL      = PatternFill("solid", fgColor="FFFBEB")
    LOW_FILL      = PatternFill("solid", fgColor="F0FDF4")
    WRAP          = Alignment(wrap_text=True, vertical="top")
    CENTER        = Alignment(horizontal="center", vertical="top")
    thin          = Side(style="thin", color="E2E8F0")
    BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_header(ws, cols: list):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.border    = BORDER
            cell.alignment = CENTER
        ws.row_dimensions[1].height = 22

    def style_row(ws, row_idx: int, fill=None):
        for cell in ws[row_idx]:
            cell.border    = BORDER
            cell.alignment = WRAP
            if fill:
                cell.fill = fill

    def set_col_widths(ws, widths: list):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def label_row(ws, label: str):
        """Write a full-width section label row."""
        ws.append([label])
        row_idx = ws.max_row
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ws.max_column or 1)
        ws[f"A{row_idx}"].fill      = SECTION_FILL
        ws[f"A{row_idx}"].font      = SECTION_FONT
        ws[f"A{row_idx}"].border    = BORDER
        ws[f"A{row_idx}"].alignment = WRAP
        ws.row_dimensions[row_idx].height = 18

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary (inserted first)
    # ══════════════════════════════════════════════════════════════════════════
    ws0 = wb.create_sheet("1. Summary")
    ws0.column_dimensions["A"].width = 26
    ws0.column_dimensions["B"].width = 60

    meta_rows = [
        ("Feature Name",    feature),
        ("Module",          r.get("detected_module", "")),
        ("Priority",        r.get("detected_priority", "")),
        ("Overall Risk",    r.get("overall_risk", "")),
        ("Complexity",      r.get("complexity_level", "")),
        ("Feature Status",  r.get("feature_status", "")),
        ("Status Reason",   r.get("feature_status_reason", "")),
        ("Total Scenarios", r.get("total_scenarios", 0)),
        ("Gherkin Cases",   len(r.get("gherkin_test_cases") or [])),
        ("Risk Areas",      len(r.get("risk_areas") or [])),
        ("Regression Items",len(r.get("regression_suite") or [])),
        ("Coverage Gaps",   len(r.get("missing_coverage") or [])),
        ("API Endpoints",   len(r.get("api_event_validation") or [])),
        ("Generated At",    r.get("generated_at", "")),
    ]
    for label, value in meta_rows:
        ws0.append([label, str(value)])
        ri = ws0.max_row
        ws0[f"A{ri}"].font      = Font(bold=True, size=10)
        ws0[f"A{ri}"].fill      = SECTION_FILL
        ws0[f"A{ri}"].border    = BORDER
        ws0[f"B{ri}"].border    = BORDER
        ws0[f"B{ri}"].alignment = WRAP
        ws0.row_dimensions[ri].height = 18

    # Feature Understanding block
    ws0.append([])
    ws0.append(["FEATURE UNDERSTANDING", r.get("feature_understanding", "")])
    ri = ws0.max_row
    ws0[f"A{ri}"].font      = Font(bold=True, size=10, color="3730A3")
    ws0[f"A{ri}"].fill      = SECTION_FILL
    ws0[f"A{ri}"].border    = BORDER
    ws0[f"B{ri}"].border    = BORDER
    ws0[f"B{ri}"].alignment = WRAP
    ws0.row_dimensions[ri].height = 100

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Test Scenarios
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("2. Test Scenarios")
    add_header(ws1, ["#", "Type", "Title", "Description", "Preconditions",
                      "Steps", "Expected Result", "Risk Level", "Traceability"])
    set_col_widths(ws1, [4, 14, 28, 40, 32, 45, 32, 10, 28])
    for i, s in enumerate(r.get("test_scenarios") or [], 1):
        risk = (s.get("risk_level") or "").lower()
        row_fill = HIGH_FILL if risk == "high" else MED_FILL if risk == "medium" else LOW_FILL if risk == "low" else None
        ws1.append([
            i,
            s.get("scenario_type") or s.get("type", ""),
            s.get("title", ""),
            s.get("description", ""),
            "\n".join(s.get("preconditions") or []),
            "\n".join(s.get("steps") or []),
            s.get("expected_result", ""),
            (s.get("risk_level") or "").upper(),
            s.get("traceability", ""),
        ])
        style_row(ws1, i + 1, fill=row_fill)
        ws1.row_dimensions[i + 1].height = 70

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Gherkin Test Cases
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("3. Gherkin Test Cases")
    add_header(ws2, ["Feature", "Scenario Title", "Tags", "Given", "When", "Then"])
    set_col_widths(ws2, [22, 32, 24, 45, 45, 45])
    for i, g in enumerate(r.get("gherkin_test_cases") or [], 1):
        ws2.append([
            g.get("feature", ""),
            g.get("scenario_title", ""),
            " ".join(g.get("tags") or []),
            "\n".join(g.get("given") or []),
            "\n".join(g.get("when") or []),
            "\n".join(g.get("then") or []),
        ])
        style_row(ws2, i + 1)
        ws2.row_dimensions[i + 1].height = 70

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Risk Areas
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("4. Risk Areas")
    add_header(ws3, ["Feature", "Module", "Priority", "Risk Score", "Reasons", "Past Bug Count"])
    set_col_widths(ws3, [28, 22, 10, 12, 56, 14])
    PRIORITY_FILLS = {"P1": P1_FILL, "P2": P2_FILL, "P3": P3_FILL, "P4": P4_FILL}
    for i, ra in enumerate(r.get("risk_areas") or [], 1):
        pf = PRIORITY_FILLS.get(ra.get("priority", ""), None)
        ws3.append([
            ra.get("feature", ""),
            ra.get("module", ""),
            ra.get("priority", ""),
            ra.get("risk_score", ""),
            "\n".join(ra.get("reasons") or []),
            ra.get("past_bug_count", 0),
        ])
        style_row(ws3, i + 1, fill=pf)
        ws3.row_dimensions[i + 1].height = 55

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — Heads-Up Warnings
    # ══════════════════════════════════════════════════════════════════════════
    ws_warn = wb.create_sheet("5. Warnings")
    add_header(ws_warn, ["#", "Warning", "Recommendation", "Severity"])
    set_col_widths(ws_warn, [4, 52, 52, 12])
    for i, w in enumerate(r.get("heads_up_warnings") or [], 1):
        ws_warn.append([
            i,
            w.get("warning", ""),
            w.get("recommendation", ""),
            w.get("severity", ""),
        ])
        style_row(ws_warn, i + 1, fill=WARNING_FILL)
        ws_warn.row_dimensions[i + 1].height = 50

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 6 — Regression Suite
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("6. Regression Suite")
    add_header(ws4, ["#", "Test Case Name", "Priority", "Module", "Reason"])
    set_col_widths(ws4, [4, 40, 12, 22, 52])
    for i, rc in enumerate(r.get("regression_suite") or [], 1):
        must = rc.get("priority", "").upper() == "MUST-RUN"
        ws4.append([
            i,
            rc.get("test_case_name", ""),
            rc.get("priority", ""),
            rc.get("module", ""),
            rc.get("reason", ""),
        ])
        style_row(ws4, i + 1, fill=P1_FILL if must else None)
        ws4.row_dimensions[i + 1].height = 40

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 7 — Coverage Gaps
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("7. Coverage Gaps")
    add_header(ws6, ["#", "Area", "Description", "Recommendation"])
    set_col_widths(ws6, [4, 26, 52, 52])
    for i, mc in enumerate(r.get("missing_coverage") or [], 1):
        ws6.append([
            i,
            mc.get("area", ""),
            mc.get("description", ""),
            mc.get("recommendation", ""),
        ])
        style_row(ws6, i + 1)
        ws6.row_dimensions[i + 1].height = 50

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 8 — API & Event Validations
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("8. API Validations")
    add_header(ws5, ["#", "Endpoint", "Method", "Validations", "Event Triggers", "DB Impacts"])
    set_col_widths(ws5, [4, 36, 10, 44, 36, 36])
    for i, api in enumerate(r.get("api_event_validation") or [], 1):
        ws5.append([
            i,
            api.get("endpoint", ""),
            api.get("method", ""),
            "\n".join(api.get("validations") or []),
            "\n".join(api.get("event_triggers") or []),
            "\n".join(api.get("db_impacts") or []),
        ])
        style_row(ws5, i + 1)
        ws5.row_dimensions[i + 1].height = 55

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 9 — Impacted Modules
    # ══════════════════════════════════════════════════════════════════════════
    ws_mod = wb.create_sheet("9. Impacted Modules")
    add_header(ws_mod, ["#", "Module ID", "Module Name", "Impact Type", "Criticality"])
    set_col_widths(ws_mod, [4, 20, 32, 20, 12])
    for i, m in enumerate(r.get("impacted_modules") or [], 1):
        ws_mod.append([
            i,
            m.get("id", ""),
            m.get("name", m.get("id", "")),
            m.get("impact_type", ""),
            m.get("criticality", ""),
        ])
        style_row(ws_mod, i + 1)
        ws_mod.row_dimensions[i + 1].height = 30

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 10 — Event Flow
    # ══════════════════════════════════════════════════════════════════════════
    ws_flow = wb.create_sheet("10. Event Flow")
    add_header(ws_flow, ["Step", "Layer", "Component", "Action", "Data", "Validation Point"])
    set_col_widths(ws_flow, [6, 16, 22, 48, 32, 40])
    LAYER_FILLS = {
        "UI":           PatternFill("solid", fgColor="EDE9FE"),
        "API":          PatternFill("solid", fgColor="DBEAFE"),
        "DB":           PatternFill("solid", fgColor="DCFCE7"),
        "Event":        PatternFill("solid", fgColor="FEF9C3"),
        "Consumer":     PatternFill("solid", fgColor="FCE7F3"),
        "Notification": PatternFill("solid", fgColor="E0F2FE"),
    }
    for i, step in enumerate(r.get("event_flow") or [], 1):
        lf = LAYER_FILLS.get(step.get("layer", ""), None)
        ws_flow.append([
            step.get("step", i),
            step.get("layer", ""),
            step.get("component", ""),
            step.get("action", ""),
            step.get("data", ""),
            step.get("validation_point", ""),
        ])
        style_row(ws_flow, i + 1, fill=lf)
        ws_flow.row_dimensions[i + 1].height = 45

    # ── Stream the workbook ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in feature)
    filename = f"QA_{safe_name}.xlsx".replace(" ", "_")

    return SR(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/projects/{project_id}/usage")
def get_project_usage(project_id: str):
    model = settings.LLM_MODEL
    price = _PRICING.get(model)
    usage = llm_client.get_project_usage(project_id)
    inp_tok = usage["input_tokens"]
    out_tok = usage["output_tokens"]
    cost = None
    if price and (inp_tok or out_tok):
        cost = {
            "input_cost": round(inp_tok / 1_000_000 * price[0], 6),
            "output_cost": round(out_tok / 1_000_000 * price[1], 6),
            "total_cost": round(inp_tok / 1_000_000 * price[0] + out_tok / 1_000_000 * price[1], 6),
        }
    return {"usage": {"input_tokens": inp_tok, "output_tokens": out_tok}, "cost": cost}


@app.post("/api/projects/{project_id}/usage/reset")
def reset_project_usage(project_id: str):
    llm_client.reset_project_usage(project_id)
    return {"status": "ok"}
