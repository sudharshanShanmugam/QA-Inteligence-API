"""Document loader – reads BRD, SRS, user stories, bug reports, API contracts, DB schemas."""

import json
import re
from pathlib import Path
from typing import Dict, Any, List, Optional
import structlog

log = structlog.get_logger()


class DocumentLoader:
    """Loads and normalises documents from various formats into a unified structure."""

    SUPPORTED_EXTENSIONS = {".txt", ".md", ".json", ".pdf", ".docx", ".sql", ".yaml", ".yml", ".xlsx", ".xls"}

    def load_text(self, content: str, doc_type: str, source_id: str, metadata: Dict[str, Any] = None) -> Dict[str, Any]:
        return {
            "source_id": source_id,
            "doc_type": doc_type,
            "content": content,
            "metadata": metadata or {},
            "format": "text",
        }

    def load_json(self, data: Dict[str, Any], doc_type: str, source_id: str) -> Dict[str, Any]:
        content = self._json_to_text(data, doc_type)
        return {
            "source_id": source_id,
            "doc_type": doc_type,
            "content": content,
            "raw": data,
            "metadata": {},
            "format": "json",
        }

    def load_file(self, file_path: str, doc_type: str) -> Dict[str, Any]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = path.suffix.lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported file type: {ext}")

        if ext == ".json":
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return self.load_json(data, doc_type, path.stem)

        elif ext == ".pdf":
            return self._load_pdf(path, doc_type)

        elif ext == ".docx":
            return self._load_docx(path, doc_type)

        elif ext in (".xlsx", ".xls"):
            return self._load_xlsx(path, doc_type)

        else:
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            return self.load_text(content, doc_type, path.stem)

    # ─── Format-specific helpers ────────────────────────────────────────────────

    def _load_pdf(self, path: Path, doc_type: str) -> Dict[str, Any]:
        try:
            import PyPDF2
            text_parts = []
            with open(path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text_parts.append(page.extract_text() or "")
            content = "\n".join(text_parts)
        except ImportError:
            content = f"[PDF parsing unavailable – install PyPDF2] File: {path.name}"
        return self.load_text(content, doc_type, path.stem)

    def _load_docx(self, path: Path, doc_type: str) -> Dict[str, Any]:
        try:
            from docx import Document
            doc = Document(path)
            content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            content = f"[DOCX parsing unavailable – install python-docx] File: {path.name}"
        return self.load_text(content, doc_type, path.stem)

    def _load_xlsx(self, path: Path, doc_type: str) -> Dict[str, Any]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    line = " | ".join(cells).strip(" |")
                    if line:
                        parts.append(line)
            content = "\n".join(parts)
        except ImportError:
            content = f"[XLSX parsing unavailable – install openpyxl] File: {path.name}"
        return self.load_text(content, doc_type, path.stem)

    def _json_to_text(self, data: Dict[str, Any], doc_type: str) -> str:
        """Convert structured JSON into prose the LLM + vector store can work with."""
        lines: List[str] = []

        if doc_type == "user_story":
            lines.append(f"USER STORY: {data.get('title', '')}")
            lines.append(f"As a {data.get('role', 'user')}, I want {data.get('goal', '')}, so that {data.get('benefit', '')}.")
            if data.get("acceptance_criteria"):
                lines.append("Acceptance Criteria:")
                for ac in data["acceptance_criteria"]:
                    lines.append(f"  - {ac}")
            if data.get("business_rules"):
                lines.append("Business Rules:")
                for br in data["business_rules"]:
                    lines.append(f"  - {br}")

        elif doc_type == "bug_report":
            lines.append(f"BUG: [{data.get('severity', 'UNKNOWN')}] {data.get('title', '')}")
            lines.append(f"Module: {data.get('module', '')} | Feature: {data.get('feature', '')}")
            lines.append(f"Description: {data.get('description', '')}")
            lines.append(f"Root Cause: {data.get('root_cause', '')}")
            lines.append(f"Steps: {data.get('steps_to_reproduce', '')}")
            lines.append(f"Status: {data.get('status', '')}")

        elif doc_type == "api_contract":
            lines.append(f"API CONTRACT: {data.get('name', '')} v{data.get('version', '1.0')}")
            for endpoint in data.get("endpoints", []):
                lines.append(f"  {endpoint.get('method', 'GET')} {endpoint.get('path', '')}: {endpoint.get('description', '')}")
                if endpoint.get("request_body"):
                    lines.append(f"    Request: {json.dumps(endpoint['request_body'])}")
                if endpoint.get("responses"):
                    lines.append(f"    Responses: {list(endpoint['responses'].keys())}")
                if endpoint.get("events_published"):
                    lines.append(f"    Publishes events: {endpoint['events_published']}")

        elif doc_type == "db_schema":
            lines.append(f"DATABASE SCHEMA: {data.get('database', '')}")
            for table in data.get("tables", []):
                lines.append(f"  TABLE: {table.get('name', '')}")
                for col in table.get("columns", []):
                    lines.append(f"    {col.get('name', '')} {col.get('type', '')} {' PK' if col.get('primary_key') else ''}{' NOT NULL' if col.get('not_null') else ''}")

        elif doc_type == "event_definition":
            lines.append(f"EVENT DEFINITION: {data.get('name', '')}")
            lines.append(f"Topic: {data.get('topic', '')} | Type: {data.get('type', '')}")
            lines.append(f"Producer: {data.get('producer', '')} → Consumer: {data.get('consumer', '')}")
            lines.append(f"Payload: {json.dumps(data.get('payload_schema', {}))}")

        else:
            lines.append(json.dumps(data, indent=2))

        return "\n".join(lines)


document_loader = DocumentLoader()
